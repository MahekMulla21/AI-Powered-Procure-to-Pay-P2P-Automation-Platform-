from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sow_config import OUTPUT_DIR, STRUCTURED_FIELDS, UNSTRUCTURED_FIELDS

THIN  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

COLORS = {
    "header_bg":  "1F3864",  # dark navy
    "header_fg":  "FFFFFF",
    "subhdr_bg":  "2E75B6",  # blue
    "alt_row":    "EBF3FF",
    "white":      "FFFFFF",
    "green":      "C6EFCE",
    "red":        "FFC7CE",
    "yellow":     "FFEB9C",
    "orange":     "FCE4D6",
    "grey":       "D9D9D9",
    "valid":      "E2EFDA",
    "missing":    "FCE4D6",
    "na":         "FFF2CC",
}


def fill(color: str):
    return PatternFill("solid", fgColor=COLORS.get(color, color))


def hdr_cell(ws, row, col, text, bg="header_bg", fg="header_fg", size=11, bold=True):
    c = ws.cell(row, col, text)
    c.font      = Font(name="Calibri", bold=bold, color=COLORS[fg], size=size)
    c.fill      = fill(bg)
    c.border    = THIN
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c


def dat_cell(ws, row, col, text, bg="white", bold=False, wrap=True,
             align="left", size=10, fg="000000"):
    c = ws.cell(row, col, str(text) if text is not None else "")
    c.font      = Font(name="Calibri", bold=bold, size=size, color=fg)
    c.fill      = fill(bg)
    c.border    = THIN
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return c


def status_bg(status: str) -> str:
    if status == "Valid":   return "valid"
    if status == "Missing": return "missing"
    return "na"


def conf_bg(cv: float) -> str:
    if cv >= 0.95: return "green"
    if cv >= 0.80: return "yellow"
    if cv >= 0.60: return "orange"
    return "red"


def get_status(val, conf):
    if val and any(x in str(val) for x in
                   ["Not Applicable", "Not Mentioned", "N/A"]):
        return "NA", "N/A"
    if not val or conf == 0.0:
        return "NA", "Missing"
    return str(val).strip(), "Valid"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — SUMMARY FORMAT
# ══════════════════════════════════════════════════════════════════════════════
def sheet_summary(wb, s_data, s_conf, u_data, u_conf):
    ws = wb.active
    ws.title = "Summary Format"

    # Title row
    ws.merge_cells("A1:D1")
    t = ws.cell(1, 1, "SOW Extraction — Summary Format")
    t.font      = Font(name="Calibri", size=14, bold=True, color="1F3864")
    t.fill      = fill("white")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    # Blank gap
    ws.row_dimensions[2].height = 6

    # Column headers
    for col, h in enumerate(
        ["Document Type", "Field Name", "Extracted Value", "Status"], 1
    ):
        hdr_cell(ws, 3, col, h, bg="subhdr_bg", size=10)
    ws.row_dimensions[3].height = 22

    # All fields in order
    SUMMARY_ROWS = [
        # label shown,         key in data,          from
        ("sow_id",             "sow_id",              "s"),
        ("reference_msa",      "reference_msa",       "s"),
        ("vendor_id",          "vendor_id",           "s"),
        ("vendor_name",        "vendor_name",         "s"),
        ("client_name",        "client_name",         "s"),
        ("project_title",      "project_title",       "s"),
        ("start_date",         "start_date",          "s"),
        ("end_date",           "end_date",            "s"),
        ("payment_terms",      "payment_terms",       "s"),
        ("currency",           "currency",            "s"),
        ("status",             "status",              "s"),
        ("total_amount",       "total_amount",        "s"),
        ("service_description","service_description", "u"),
        ("scope_of_work",      "scope_of_work",       "u"),
        ("deliverables",       "deliverables",        "u"),
        ("payment_schedule",   "payment_schedule",    "u"),
        ("resource_requirements","resource_requirements","u"),
        ("acceptance_criteria","acceptance_criteria", "u"),
        ("termination_clause", "termination_clause",  "u"),
    ]

    for idx, (label, key, src) in enumerate(SUMMARY_ROWS):
        row  = 4 + idx
        val  = s_data.get(key) if src == "s" else u_data.get(key)
        conf = s_conf.get(key, 0.0) if src == "s" else u_conf.get(key, 0.0)
        fv, fs = get_status(val, conf)

        # Row height based on content lines
        n_lines = fv.count("\n") + 1 if fv != "NA" else 1
        ws.row_dimensions[row].height = max(20, min(120, n_lines * 15 + 6))

        bg_row = "alt_row" if idx % 2 == 0 else "white"
        dat_cell(ws, row, 1, "SOW", bg_row, bold=True, align="center")
        dat_cell(ws, row, 2, label, bg_row, bold=True)
        dat_cell(ws, row, 3, fv, bg_row, wrap=True)
        dat_cell(ws, row, 4, fs, status_bg(fs), align="center", bold=True)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 12


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — STRUCTURED FIELDS
# ══════════════════════════════════════════════════════════════════════════════
def sheet_structured(wb, data, conf):
    ws = wb.create_sheet("Structured Fields")

    ws.merge_cells("A1:G1")
    hdr_cell(ws, 1, 1, f"STRUCTURED FIELDS  ({len(STRUCTURED_FIELDS)} Fields)", size=12)
    ws.row_dimensions[1].height = 32

    headers = ["#", "Field Name", "Extracted Value", "Confidence", "Method", "Status"]
    for col, h in enumerate(headers, 1):
        hdr_cell(ws, 2, col, h, bg="subhdr_bg", size=10)
    ws.row_dimensions[2].height = 22

    for idx, field in enumerate(STRUCTURED_FIELDS, 1):
        row  = idx + 2
        val  = data.get(field)
        cv   = conf.get(field, 0.0)
        bg   = "alt_row" if idx % 2 == 0 else "white"
        st   = "✔  Found" if val else "✘  Missing"
        sbg  = "green" if val else "red"

        dat_cell(ws, row, 1, idx, bg, align="center")
        dat_cell(ws, row, 2, field, bg, bold=True)
        dat_cell(ws, row, 3, str(val) if val else "—", bg, wrap=True)
        dat_cell(ws, row, 4, f"{cv*100:.0f}%", conf_bg(cv), align="center", bold=True)
        dat_cell(ws, row, 5, "Dictionary label lookup", bg)
        dat_cell(ws, row, 6, st, sbg, align="center", bold=True)
        ws.row_dimensions[row].height = 22

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 14


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — UNSTRUCTURED FIELDS
# ══════════════════════════════════════════════════════════════════════════════
def sheet_unstructured(wb, data, conf):
    ws = wb.create_sheet("Unstructured Fields")

    ws.merge_cells("A1:F1")
    hdr_cell(ws, 1, 1, f"UNSTRUCTURED FIELDS  ({len(UNSTRUCTURED_FIELDS)} Fields)", size=12)
    ws.row_dimensions[1].height = 32

    headers = ["#", "Field Name", "Extracted Value", "Confidence", "Method", "Status"]
    for col, h in enumerate(headers, 1):
        hdr_cell(ws, 2, col, h, bg="subhdr_bg", size=10)
    ws.row_dimensions[2].height = 22

    for idx, field in enumerate(UNSTRUCTURED_FIELDS, 1):
        row  = idx + 2
        val  = data.get(field)
        cv   = conf.get(field, 0.0)
        bg   = "alt_row" if idx % 2 == 0 else "white"
        flat = str(val).strip() if val else "—"
        st   = "✔  Found" if val else "✘  Missing"
        sbg  = "green" if val else "red"
        n    = flat.count("\n") + 1

        dat_cell(ws, row, 1, idx, bg, align="center")
        dat_cell(ws, row, 2, field, bg, bold=True)
        dat_cell(ws, row, 3, flat, bg, wrap=True)
        dat_cell(ws, row, 4, f"{cv*100:.0f}%", conf_bg(cv), align="center", bold=True)
        dat_cell(ws, row, 5, "Section parsing", bg)
        dat_cell(ws, row, 6, st, sbg, align="center", bold=True)
        ws.row_dimensions[row].height = max(40, min(200, n * 15 + 10))

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 14


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — ACCURACY REPORT
# ══════════════════════════════════════════════════════════════════════════════
def sheet_accuracy(wb, s_data, s_conf, u_data, u_conf, s_found, u_found):
    ws = wb.create_sheet("Accuracy Report")

    ws.merge_cells("A1:D1")
    hdr_cell(ws, 1, 1, "ACCURACY & CONFIDENCE REPORT", size=12)
    ws.row_dimensions[1].height = 32

    total  = len(STRUCTURED_FIELDS) + len(UNSTRUCTURED_FIELDS)
    found  = s_found + u_found
    rate   = round(found / total * 100, 1) if total else 0
    all_cv = list(s_conf.values()) + list(u_conf.values())
    avg_cv = round(sum(all_cv) / len(all_cv) * 100, 1) if all_cv else 0

    # KPI section
    hdr_cell(ws, 3, 1, "KPI", bg="subhdr_bg"); hdr_cell(ws, 3, 2, "Value", bg="subhdr_bg")
    ws.row_dimensions[3].height = 20

    kpis = [
        ("Total Fields Targeted", total),
        ("Fields Found",          found),
        ("Fields Missing",        total - found),
        ("Extraction Rate",       f"{rate}%"),
        ("Avg Confidence",        f"{avg_cv}%"),
        (f"Structured ({len(STRUCTURED_FIELDS)} fields)",  f"{s_found}/{len(STRUCTURED_FIELDS)} found"),
        (f"Unstructured ({len(UNSTRUCTURED_FIELDS)} fields)", f"{u_found}/{len(UNSTRUCTURED_FIELDS)} found"),
    ]
    for r_off, (k, v) in enumerate(kpis, 4):
        bg = "alt_row" if r_off % 2 == 0 else "white"
        dat_cell(ws, r_off, 1, k, bg, bold=True)
        dat_cell(ws, r_off, 2, str(v), bg)
        ws.row_dimensions[r_off].height = 20

    # Per-field breakdown
    rs = len(kpis) + 6
    ws.merge_cells(f"A{rs}:D{rs}")
    hdr_cell(ws, rs, 1, "FIELD-BY-FIELD BREAKDOWN", bg="subhdr_bg")
    ws.row_dimensions[rs].height = 20
    rs += 1

    for col, h in enumerate(["Field Name", "Type", "Confidence", "Status"], 1):
        hdr_cell(ws, rs, col, h, bg="subhdr_bg", size=10)
    ws.row_dimensions[rs].height = 20
    rs += 1

    all_fields = (
        [(f, "Structured", s_conf.get(f, 0.0), s_data.get(f)) for f in STRUCTURED_FIELDS] +
        [(f, "Unstructured", u_conf.get(f, 0.0), u_data.get(f)) for f in UNSTRUCTURED_FIELDS]
    )
    for r_off, (field, ftype, cv, val) in enumerate(all_fields):
        row = rs + r_off
        bg  = "alt_row" if r_off % 2 == 0 else "white"
        st  = "✔  Found" if val else "✘  Missing"
        sbg = "green" if val else "red"
        dat_cell(ws, row, 1, field, bg, bold=True)
        dat_cell(ws, row, 2, ftype, bg, align="center")
        dat_cell(ws, row, 3, f"{cv*100:.0f}%", conf_bg(cv), align="center", bold=True)
        dat_cell(ws, row, 4, st, sbg, align="center", bold=True)
        ws.row_dimensions[row].height = 20

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC FUNCTION
# ══════════════════════════════════════════════════════════════════════════════
def save_excel(s_data, s_conf, u_data, u_conf,
               s_found, u_found, output_filename) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, output_filename)

    wb = Workbook()

    sheet_summary(wb, s_data, s_conf, u_data, u_conf)
    sheet_structured(wb, s_data, s_conf)
    sheet_unstructured(wb, u_data, u_conf)
    sheet_accuracy(wb, s_data, s_conf, u_data, u_conf, s_found, u_found)

    wb.save(out_path)
    return out_path
